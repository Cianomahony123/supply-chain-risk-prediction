"""
export_for_powerbi.py

Reads the three source CSVs, applies the same fusion/standardisation pipeline
used in FP_JNB.ipynb, then exports Power BI-ready tables into a single workbook:
powerbi/SupplyChainRisk_PowerBI.xlsx

Run from the repo root:
    python powerbi/export_for_powerbi.py
"""

import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = os.path.join(ROOT, "powerbi", "SupplyChainRisk_PowerBI.xlsx")

LOGISTICS_CSV = os.path.join(ROOT, "global_supply_chain_risk_2026.csv")
FUSED_CSV     = os.path.join(ROOT, "final_supply_chain_project_data.csv")
GSCPI_CSV     = os.path.join(ROOT, "gscpi_data.csv")


# ---------------------------------------------------------------------------
# Load & merge
# ---------------------------------------------------------------------------

def load_data():
    logistics = pd.read_csv(LOGISTICS_CSV, parse_dates=["Date"])
    fused     = pd.read_csv(FUSED_CSV,     parse_dates=["Date"])
    gscpi     = pd.read_csv(GSCPI_CSV)

    # Merge Disruption_Occurred from logistics into fused (on Shipment_ID)
    if "Disruption_Occurred" not in fused.columns:
        fused = fused.merge(
            logistics[["Shipment_ID", "Disruption_Occurred",
                        "Fuel_Price_Index", "Geopolitical_Risk_Score",
                        "Weather_Condition", "Carrier_Reliability_Score"]],
            on="Shipment_ID", how="left"
        )
    return logistics, fused, gscpi


# ---------------------------------------------------------------------------
# Table 1: Shipments  (main fact table)
# ---------------------------------------------------------------------------

def build_shipments(logistics):
    df = logistics.copy()
    df["Date"] = pd.to_datetime(df["Date"])
    df["Year"]  = df["Date"].dt.year
    df["Month"] = df["Date"].dt.month
    df["Month_Name"] = df["Date"].dt.strftime("%b %Y")
    df["Disruption_Label"] = df["Disruption_Occurred"].map({1: "Disrupted", 0: "On-Time"})

    # Risk tier based on geopolitical score
    df["Risk_Tier"] = pd.cut(
        df["Geopolitical_Risk_Score"],
        bins=[0, 3, 6, 10],
        labels=["Low", "Medium", "High"],
        right=True
    )
    return df[[
        "Shipment_ID", "Date", "Year", "Month", "Month_Name",
        "Origin_Port", "Destination_Port", "Transport_Mode", "Product_Category",
        "Distance_km", "Weight_MT", "Lead_Time_Days",
        "Fuel_Price_Index", "Geopolitical_Risk_Score", "Risk_Tier",
        "Weather_Condition", "Carrier_Reliability_Score",
        "Disruption_Occurred", "Disruption_Label"
    ]]


# ---------------------------------------------------------------------------
# Table 2: Disruption_by_Mode  (bar chart — transport mode breakdown)
# ---------------------------------------------------------------------------

def build_by_mode(logistics):
    df = logistics.groupby("Transport_Mode").agg(
        total_shipments=("Shipment_ID", "count"),
        disruptions=("Disruption_Occurred", "sum"),
        avg_lead_time=("Lead_Time_Days", "mean"),
        avg_distance_km=("Distance_km", "mean"),
    ).reset_index()
    df["disruption_rate_pct"] = (df["disruptions"] / df["total_shipments"] * 100).round(1)
    df["avg_lead_time"] = df["avg_lead_time"].round(1)
    df["avg_distance_km"] = df["avg_distance_km"].round(0)
    return df.sort_values("disruption_rate_pct", ascending=False)


# ---------------------------------------------------------------------------
# Table 3: Disruption_by_Category  (bar chart — product category breakdown)
# ---------------------------------------------------------------------------

def build_by_category(logistics):
    df = logistics.groupby("Product_Category").agg(
        total_shipments=("Shipment_ID", "count"),
        disruptions=("Disruption_Occurred", "sum"),
        avg_lead_time=("Lead_Time_Days", "mean"),
    ).reset_index()
    df["disruption_rate_pct"] = (df["disruptions"] / df["total_shipments"] * 100).round(1)
    df["avg_lead_time"] = df["avg_lead_time"].round(1)
    return df.sort_values("disruption_rate_pct", ascending=False)


# ---------------------------------------------------------------------------
# Table 4: Monthly_Trends  (line chart — disruptions over time)
# ---------------------------------------------------------------------------

def build_monthly_trends(logistics):
    logistics["YearMonth"] = pd.to_datetime(logistics["Date"]).dt.to_period("M").dt.to_timestamp()
    df = logistics.groupby("YearMonth").agg(
        total_shipments=("Shipment_ID", "count"),
        disruptions=("Disruption_Occurred", "sum"),
        avg_lead_time=("Lead_Time_Days", "mean"),
        avg_gscpi=("Fuel_Price_Index", "mean"),  # proxy for macro stress
    ).reset_index()
    df["disruption_rate_pct"] = (df["disruptions"] / df["total_shipments"] * 100).round(1)
    df["avg_lead_time"] = df["avg_lead_time"].round(1)
    return df


# ---------------------------------------------------------------------------
# Table 5: Route_Summary  (map visual — origin/destination pairs)
# ---------------------------------------------------------------------------

def build_route_summary(logistics):
    df = logistics.groupby(["Origin_Port", "Destination_Port"]).agg(
        total_shipments=("Shipment_ID", "count"),
        disruptions=("Disruption_Occurred", "sum"),
        avg_lead_time=("Lead_Time_Days", "mean"),
        avg_distance_km=("Distance_km", "mean"),
    ).reset_index()
    df["disruption_rate_pct"] = (df["disruptions"] / df["total_shipments"] * 100).round(1)

    # Port coordinates for Power BI map visual
    PORT_COORDS = {
        "Rotterdam":    (51.9225, 4.4792),
        "Shanghai":     (31.2304, 121.4737),
        "Singapore":    (1.3521,  103.8198),
        "Los Angeles":  (33.7285, -118.2620),
        "New York":     (40.6643, -74.0060),
        "Dubai":        (25.2697,  55.3095),
        "Busan":        (35.1796, 129.0756),
        "Hamburg":      (53.5753,  10.0153),
        "Mumbai":       (19.0760,  72.8777),
        "Tokyo":        (35.6762, 139.6503),
    }
    df["origin_lat"]  = df["Origin_Port"].map(lambda p: PORT_COORDS.get(p, (np.nan, np.nan))[0])
    df["origin_lon"]  = df["Origin_Port"].map(lambda p: PORT_COORDS.get(p, (np.nan, np.nan))[1])
    df["dest_lat"]    = df["Destination_Port"].map(lambda p: PORT_COORDS.get(p, (np.nan, np.nan))[0])
    df["dest_lon"]    = df["Destination_Port"].map(lambda p: PORT_COORDS.get(p, (np.nan, np.nan))[1])
    return df.sort_values("disruption_rate_pct", ascending=False)


# ---------------------------------------------------------------------------
# Table 6: Model_Comparison  (table visual — SVM vs MLP vs RF)
# ---------------------------------------------------------------------------

def build_model_comparison():
    return pd.DataFrame([
        {
            "Model":           "Balanced SVM (RBF)",
            "Accuracy_%":      61.4,
            "Recall_Minority": 0.74,
            "Precision":       0.33,
            "F1_Score":        0.45,
            "Recommended":     "Yes",
            "Reason":          "Catches 74% of actual disruptions — operationally superior despite lower accuracy"
        },
        {
            "Model":           "MLP Neural Network",
            "Accuracy_%":      72.6,
            "Recall_Minority": 0.20,
            "Precision":       0.31,
            "F1_Score":        0.24,
            "Recommended":     "No",
            "Reason":          "High accuracy is majority-class bias — misses 80% of disruptions"
        },
        {
            "Model":           "Random Forest",
            "Accuracy_%":      71.1,
            "Recall_Minority": 0.18,
            "Precision":       None,
            "F1_Score":        None,
            "Recommended":     "No",
            "Reason":          "Same majority-class bias as MLP"
        },
    ])


# ---------------------------------------------------------------------------
# Table 7: Risk_Scatter  (fused data for scatter plot)
# ---------------------------------------------------------------------------

def build_risk_scatter(fused):
    df = fused[["Shipment_ID", "Lead_Time_Days", "GSCPI",
                "gust_kph", "precip_mm", "Transport_Mode",
                "Product_Category", "Distance_km"]].copy()
    # Pull disruption flag from fused if available
    if "Disruption_Occurred" in fused.columns:
        df["Disruption_Occurred"] = fused["Disruption_Occurred"]
        df["Disruption_Label"] = df["Disruption_Occurred"].map({1: "Disrupted", 0: "On-Time"})
    df = df.rename(columns={
        "Lead_Time_Days": "Lead_Time_Days_Zscore",
        "GSCPI": "GSCPI_Zscore",
        "gust_kph": "Gust_Zscore",
        "precip_mm": "Precip_Zscore"
    })
    return df


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_sheet(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def write_excel(sheets: dict):
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl", datetime_format="YYYY-MM-DD") as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name, index=False)
    wb = load_workbook(OUT_PATH)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    logistics, fused, gscpi = load_data()

    sheets = {
        "Shipments":          build_shipments(logistics),
        "By_Transport_Mode":  build_by_mode(logistics),
        "By_Product_Category": build_by_category(logistics),
        "Monthly_Trends":     build_monthly_trends(logistics),
        "Route_Summary":      build_route_summary(logistics),
        "Model_Comparison":   build_model_comparison(),
        "Risk_Scatter":       build_risk_scatter(fused),
    }

    write_excel(sheets)

    print("\nSheets written:")
    for name, df in sheets.items():
        print(f"  {name}: {len(df)} rows x {len(df.columns)} cols")


if __name__ == "__main__":
    main()
